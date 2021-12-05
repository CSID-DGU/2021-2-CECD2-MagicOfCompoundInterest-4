from flask import Flask, request
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import create_engine
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import scoped_session, sessionmaker
from pathlib import Path
from openpyxl import Workbook, load_workbook
from sqlalchemy.sql.expression import column

from config import DATABASE_URI

app = Flask(__name__)
engine = create_engine(DATABASE_URI, convert_unicode=True)
db_session = scoped_session(sessionmaker(autocommit=False, autoflush=False, bind=engine))

Base = declarative_base()
Base.query = db_session.query_property()

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URI
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = True
app.config['SQLALCHEMY_ECHO'] = True
app.config['SECRET_KEY'] = 'super secret key'
app.config['SESSION_TYPE'] = 'filesystem'

Base = SQLAlchemy(app)

from model import *

Base.create_all()

@app.route('/', methods=['GET'])
def get_list():
    datetime = request.args.get('date')
    result = Welfare.query.filter_by(created_at=datetime).all()

    if len(result) == 0:
        return 'Empty', 404

    return result

@app.route('/', methods=['POST'])
def create_welfare():
    file = request.files['attachment'].read()
    wb = load_workbook(file)
    sheet = wb.get_sheet_by_name('result20211125')
    return sheet.cell(row=1, column=1).value

if __name__ == '__main__':
    app.run()